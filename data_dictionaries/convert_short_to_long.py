#!/usr/bin/env python3
"""
Convert legacy cell line JSON records from short values to long values.

Reads the short->long mapping from the ASCR data dictionary Excel file,
then recursively converts all matching field values in the input JSON files.

Usage:
    # Convert all JSON files in old_data/ to old_data_converted/
    python data_dictionaries/convert_short_to_long.py old_data/ old_data_converted/

    # Convert a single file
    python data_dictionaries/convert_short_to_long.py old_data/FINi003-A.json old_data_converted/
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path("data_dictionaries/2025_12_ascr_data_dictionary_v1.0.xlsx")


def build_mapping(xlsx_path: Path) -> dict:
    """Build {field_name: {short_value: long_value}} from the data dictionary."""
    wb = load_workbook(xlsx_path)
    ws = wb["data_dictionary"]

    def parse_values(val):
        if not val or str(val).strip() in ("NA", "None", ""):
            return []
        val = str(val).strip()
        if val.startswith("[") and val.endswith("]"):
            val = val[1:-1]
        return [v.strip() for v in re.split(r"[,\n]", val) if v.strip()]

    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        field = row[1]
        short_raw = row[11]
        long_raw = row[12]

        if not field or not short_raw or str(short_raw).strip() in ("NA", "None", ""):
            continue

        # Skip ontology URLs and boolean fields (TRUE/FALSE don't need conversion)
        short_str = str(short_raw).strip()
        if "purl.obolibrary" in short_str or short_str in ("[TRUE,FALSE]", "TRUE,FALSE"):
            continue

        shorts = parse_values(short_raw)
        longs = parse_values(long_raw)

        if len(shorts) != len(longs):
            print(f"  Warning: skipping {field} — mismatched value counts ({len(shorts)} vs {len(longs)})")
            continue

        if field not in mapping:
            mapping[field] = {}

        for s, l in zip(shorts, longs):
            if s != l:
                mapping[field][s] = l

    # JSON field names that differ from the Excel field names
    FIELD_ALIASES = {
        'non_int_vector_type': 'non_int_vector',  # JSON uses _type suffix
        'germ_layer': 'cell_type',                # JSON uses germ_layer, Excel uses cell_type (2nd occurrence)
    }
    for json_field, excel_field in FIELD_ALIASES.items():
        if excel_field in mapping:
            mapping[json_field] = mapping[excel_field]

    return mapping


def convert_value(key: str, value, mapping: dict):
    """Convert a single field value using the mapping."""
    if isinstance(value, str) and key in mapping:
        return mapping[key].get(value, value)
    return value


def convert_record(data, mapping: dict, parent_key: str = None):
    """Recursively convert all field values in a JSON object."""
    if isinstance(data, dict):
        return {k: convert_record(v, mapping, parent_key=k) for k, v in data.items()}
    elif isinstance(data, list):
        if parent_key and parent_key in mapping:
            # List of strings (e.g. genotype: ["PC", "GC"])
            return [
                mapping[parent_key].get(item, item) if isinstance(item, str) else convert_record(item, mapping)
                for item in data
            ]
        else:
            return [convert_record(item, mapping) for item in data]
    elif isinstance(data, str):
        if data == '':
            return None
        if parent_key and parent_key in mapping:
            return mapping[parent_key].get(data, data)
    return data


def convert_file(input_path: Path, output_path: Path, mapping: dict) -> int:
    """Convert one JSON file. Returns number of values changed."""
    with open(input_path) as f:
        original = json.load(f)

    converted = convert_record(original, mapping)

    original_str = json.dumps(original, sort_keys=True)
    converted_str = json.dumps(converted, sort_keys=True)
    changes = sum(1 for a, b in zip(original_str, converted_str) if a != b)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(converted, f, indent=2)

    return 0 if original_str == converted_str else 1


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    print(f"Loading mapping from {XLSX_PATH}...")
    mapping = build_mapping(XLSX_PATH)
    mapped_fields = sum(len(v) for v in mapping.values())
    print(f"  {len(mapping)} fields, {mapped_fields} short->long pairs\n")

    # Collect input files
    if input_path.is_dir():
        input_files = sorted(input_path.glob("*.json"))
    elif input_path.is_file():
        input_files = [input_path]
    else:
        print(f"Error: {input_path} not found")
        sys.exit(1)

    if not input_files:
        print(f"No JSON files found in {input_path}")
        sys.exit(1)

    changed = 0
    unchanged = 0
    for f in input_files:
        if output_path.suffix == ".json":
            out = output_path
        else:
            out = output_path / f.name

        was_changed = convert_file(f, out, mapping)
        if was_changed:
            print(f"  converted  {f.name}")
            changed += 1
        else:
            print(f"  unchanged  {f.name}")
            unchanged += 1

    print(f"\nDone: {changed} converted, {unchanged} unchanged")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()
