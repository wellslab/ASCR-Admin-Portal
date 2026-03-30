import json
import os
import time
from typing import List, Dict, Any, Tuple, Optional
import asyncio
import io
import typing
from dataclasses import dataclass
from pathlib import Path
from openai import AsyncOpenAI
import logging
from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError
from agents import Agent, Runner
from data_dictionaries.models import JSONOutputSchema
from config_manager import config_manager

_AI_ASSETS = Path(__file__).parent / "ai_assets"

logger = logging.getLogger(__name__)

@dataclass
class PDFInfo:
    file_id: str
    filename: str
    client: AsyncOpenAI


def _literal_values(annotation) -> list | None:
    """Return Literal values if annotation is/wraps a Literal, else None."""
    origin = typing.get_origin(annotation)
    args = typing.get_args(annotation)
    if origin is typing.Literal:
        return list(args)
    if origin is typing.Union:  # Optional[Literal[...]]
        for arg in args:
            if typing.get_origin(arg) is typing.Literal:
                return list(typing.get_args(arg))
    if origin is list and args:  # List[Literal[...]]
        if typing.get_origin(args[0]) is typing.Literal:
            return list(typing.get_args(args[0]))
    return None


def _nested_model(annotation) -> type | None:
    """Return the Pydantic BaseModel class buried in an annotation, if any."""
    origin = typing.get_origin(annotation)
    args = typing.get_args(annotation)
    if isinstance(annotation, type) and issubclass(annotation, BaseModel):
        return annotation
    if origin is typing.Union:
        for arg in args:
            if isinstance(arg, type) and issubclass(arg, BaseModel):
                return arg
    if origin is list and args:
        if isinstance(args[0], type) and issubclass(args[0], BaseModel):
            return args[0]
    return None


def load_llm_curation_instructions() -> str:
    """Read per-field curation instructions from the data dictionary Excel file.
    Returns a markdown string for fields where llm_curate = YES.
    Path is read from DATA_DICT_PATH in config (Settings page).
    """
    raw_path = config_manager.get("DATA_DICT_PATH")
    if not raw_path:
        raise FileNotFoundError("DATA_DICT_PATH is not configured. Set it in Settings.")
    path = Path(raw_path)
    if not path.is_absolute():
        path = Path(__file__).parent / path
    if not path.exists():
        raise FileNotFoundError(f"Data dictionary file not found: {path}. Update DATA_DICT_PATH in Settings.")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["data_dictionary"]

    by_class: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        class_name, field_name, key, llm_curate, llm_instructions = (
            row[0], row[1], row[5], row[13], row[14]
        )
        if not class_name or not field_name:
            continue
        if key and str(key).upper() in ("PK", "FK"):
            continue
        if not llm_curate or str(llm_curate).upper() != "YES":
            continue
        if not llm_instructions:
            continue
        by_class.setdefault(class_name, []).append((field_name, llm_instructions))
    wb.close()

    lines = [
        "# LLM Curation Instructions\n",
        "Fields marked for LLM curation in the data dictionary.\n",
        "Use `null` for any field where the value is unknown, not reported, or not found in the article.\n",
    ]
    for class_name, fields in by_class.items():
        lines += [f"\n## {class_name}\n", "| Field | Instruction |", "|-------|-------------|"]
        lines += [f"| `{f}` | {instr} |" for f, instr in fields]
    return "\n".join(lines)


def get_enum_constraints_json() -> str:
    """
    Extract all Literal field constraints from JSONOutputSchema and nested models.
    Returns a JSON string: {"ModelName.field_name": ["allowed", "values"], ...}
    Used to provide the normalisation agent with up-to-date controlled vocabularies.
    """
    def _collect(model_class: type, visited: set) -> dict:
        if model_class in visited:
            return {}
        visited.add(model_class)
        result = {}
        for field_name, field_info in model_class.model_fields.items():
            annotation = field_info.annotation
            literals = _literal_values(annotation)
            if literals is not None:
                result[f"{model_class.__name__}.{field_name}"] = literals
            else:
                nested = _nested_model(annotation)
                if nested:
                    result.update(_collect(nested, visited))
        return result

    constraints = _collect(JSONOutputSchema, set())
    return json.dumps(constraints, indent=2)

async def validate_and_upload_pdf(filename: str, file_data: bytes) -> PDFInfo:
    """
    Validate PDF file and upload to OpenAI Files API.
    
    Args:
        filename: Name of the uploaded PDF file
        file_data: PDF bytes
        
    Returns:
        PDFInfo containing file_id, filename, and client
    """
    logger.info(f"Validating file: {filename}")
    
    # Validate PDF file
    if not filename.endswith(".pdf"):
        error_msg = "File must be a PDF, please try again."
        logger.error(f"Validation failed: {error_msg}")
        raise ValueError(error_msg)
    
    # Initialize OpenAI client with API key from config
    api_key = config_manager.get("OPENAI_API_KEY")
    if not api_key or api_key == "your_openai_api_key_here":
        raise ValueError("OpenAI API key not set. Please set it in Settings.")

    # Log masked API key for debugging
    masked_key = f"***{api_key[-4:]}" if api_key and len(api_key) > 4 else "***"
    logger.info(f"Using OpenAI API key ending with: {masked_key}")

    client = AsyncOpenAI(api_key=api_key)
    
    # Upload PDF file to OpenAI Files API
    logger.info(f"Uploading PDF file {filename} to OpenAI Files API...")
    file_obj = io.BytesIO(file_data)
    file_obj.name = filename
    
    pdf = await client.files.create(
        file=file_obj,
        purpose="user_data"
    )
    logger.info(f"PDF uploaded successfully with ID: {pdf.id}")
    
    return PDFInfo(file_id=pdf.id, filename=filename, client=client)

def start_identification_agent():
    with open(_AI_ASSETS / "identification_prompt.md", "r") as f:
        prompt = f.read()

    model = config_manager.get("SELECTED_MODEL", "gpt-4.1-mini")

    CellLineIdentificationAgent = Agent(
        name="CellLineIdentificationAgent",
        tools=[],
        model=model,
        instructions=prompt,
        output_type=List[str]
    )
    return CellLineIdentificationAgent

def start_curation_agent():
    with open(_AI_ASSETS / "curation_prompt.md", "r") as f:
        cell_line_curation_prompt = f.read()

    llm_curation_instructions = load_llm_curation_instructions()
    curation_prompt_combined = cell_line_curation_prompt + '\n\n' + llm_curation_instructions

    model = config_manager.get("SELECTED_MODEL", "gpt-4.1-mini")

    CellLineCurationAgent = Agent(
        name="CellLineCurationAgent",
        tools=[],
        model=model,
        instructions=curation_prompt_combined,
        output_type=JSONOutputSchema
    )

    return CellLineCurationAgent

def start_normalisation_agent():
    with open(_AI_ASSETS / "normalisation_prompt.md", "r") as f:
        prompt = f.read()

    model = config_manager.get("SELECTED_MODEL", "gpt-4.1-mini")

    CellLineNormalisationAgent = Agent(
        name="CellLineNormalisationAgent",
        tools=[],
        model=model,
        instructions=prompt,
        output_type=JSONOutputSchema
    )
    return CellLineNormalisationAgent

def initialize_agents() -> Tuple[Any, Any, Any]:
    """
    Initialize the three AI agents for the curation pipeline.
    
    Returns:
        Tuple of (identification_agent, curation_agent, normalization_agent)
    """
    logger.info("Initializing AI agents...")
    
    identification_agent = start_identification_agent()
    curation_agent = start_curation_agent()
    normalization_agent = start_normalisation_agent()
    
    logger.info("All agents initialized successfully")
    return identification_agent, curation_agent, normalization_agent

async def identify_cell_lines(pdf_info: PDFInfo, identification_agent: Any) -> List[str]:
    """
    Run cell line identification on the PDF using the identification agent.
    
    Args:
        pdf_info: PDF information including file_id
        identification_agent: Agent for identifying cell lines
        
    Returns:
        List of identified cell line IDs
    """
    logger.info("STAGE 1: Running Cell Line Identification Agent...")
    identification_start = time.time()
    
    # Prepare input for agent
    pdf_input = [
        {
            "role": "user",
            "content": [
                {
                    "type": "input_file",
                    "file_id": pdf_info.file_id,
                }
            ],
        }
    ]
    
    try:
        identification_result = await asyncio.to_thread(
            Runner.run_sync,
            identification_agent,
            pdf_input
        )
        identification_time = time.time() - identification_start
        
        logger.info(f"Identification completed in {identification_time:.2f}s")
        logger.info(f"Identification result: {identification_result}")
        
        # Extract final output from RunResult object
        cell_lines_found = identification_result.final_output if identification_result.final_output else []
        
        if not cell_lines_found:
            error_msg = "No cell lines identified in the document"
            logger.warning(f"{error_msg}")
            raise ValueError(error_msg)
        
        logger.info(f"Found {len(cell_lines_found)} cell lines: {cell_lines_found}")
        return cell_lines_found
        
    except Exception as e:
        error_msg = f"Identification stage failed: {str(e)}"
        logger.error(f"{error_msg}", exc_info=True)
        raise Exception(error_msg)

async def _curate_one(pdf_info: PDFInfo, curation_agent: Any, cell_line_id: str) -> Optional[Dict[str, Any]]:
    """Curate a single cell line. Returns result dict or None on failure."""
    curation_input = [
        {
            "role": "user",
            "content": [
                {"type": "input_file", "file_id": pdf_info.file_id},
                {"type": "input_text", "text": f"For the cell line named {cell_line_id}, run metadata curation on the given file using your instructions."},
            ],
        }
    ]
    start = time.time()
    try:
        result = await asyncio.to_thread(Runner.run_sync, curation_agent, curation_input)
        curation_data = result.final_output.model_dump() if result and result.final_output else None
        if curation_data:
            return {"cell_line_id": cell_line_id, "curation_data": curation_data, "curation_time": time.time() - start}
        raise ValueError(f"No curation result returned for {cell_line_id}")
    except Exception as e:
        logger.error(f"Curation failed for {cell_line_id}: {e}", exc_info=True)
        raise


async def _normalize_one(normalization_agent: Any, cell_line_id: str, curated: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Normalize a single curated cell line. Returns result dict or None on failure."""
    curation_data = curated["curation_data"]
    metadata_list = curation_data if isinstance(curation_data, list) else [curation_data]
    start = time.time()
    try:
        # Use first metadata object (standard case)
        metadata_obj = metadata_list[0]
        enum_constraints = get_enum_constraints_json()
        normalization_input = [
            {
                "role": "user",
                "content": [{"type": "input_text", "text": f"Controlled vocabulary constraints:\n{enum_constraints}\n\nNormalize this cell line metadata for {cell_line_id}: {metadata_obj}"}],
            }
        ]
        result = await asyncio.to_thread(Runner.run_sync, normalization_agent, normalization_input)
        normalized_data = result.final_output.model_dump() if result and result.final_output else None
        if normalized_data:
            return {
                "cell_line_id": cell_line_id,
                "metadata_object_index": 0,
                "normalized_data": normalized_data,
                "processing_times": {
                    "curation_seconds": curated["curation_time"],
                    "normalization_seconds": time.time() - start,
                },
            }
        raise ValueError(f"No normalization result returned for {cell_line_id}")
    except Exception as e:
        logger.error(f"Normalization failed for {cell_line_id}: {e}", exc_info=True)
        raise


async def process_single_cell_line(
    pdf_info: PDFInfo,
    curation_agent: Any,
    normalization_agent: Any,
    cell_line_id: str,
    progress_cb,
) -> Optional[Dict[str, Any]]:
    """Full pipeline for one cell line: curate → normalize. Calls progress_cb(name, stage, status, error_message) at each transition."""
    try:
        progress_cb(cell_line_id, "curating", "processing")
        try:
            curated = await _curate_one(pdf_info, curation_agent, cell_line_id)
        except Exception as e:
            progress_cb(cell_line_id, "curating", "failed", str(e))
            return None
        progress_cb(cell_line_id, "curating", "completed")

        progress_cb(cell_line_id, "normalizing", "processing")
        try:
            normalized = await _normalize_one(normalization_agent, cell_line_id, curated)
        except Exception as e:
            progress_cb(cell_line_id, "normalizing", "failed", str(e))
            return None
        progress_cb(cell_line_id, "normalizing", "completed")

        return normalized
    except Exception as e:
        logger.error(f"Pipeline failed for {cell_line_id}: {e}", exc_info=True)
        return None


async def run_parallel_pipeline(
    pdf_info: PDFInfo,
    curation_agent: Any,
    normalization_agent: Any,
    cell_lines: List[str],
    progress_cb,
) -> List[Dict[str, Any]]:
    """Run curate+normalize for all cell lines concurrently via asyncio.gather."""
    tasks = [
        process_single_cell_line(pdf_info, curation_agent, normalization_agent, cl, progress_cb)
        for cl in cell_lines
    ]
    results = await asyncio.gather(*tasks, return_exceptions=True)
    return [r for r in results if r is not None and not isinstance(r, Exception)]



async def validate_cell_lines(normalized_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Validate normalized cell line data against JSONOutputSchema Pydantic model.

    Args:
        normalized_results: List of normalized results from previous stage

    Returns:
        List of validated results with validation status
    """
    logger.info("STAGE 4: Validating cell line data against JSONOutputSchema...")

    validated_results = []
    validation_errors = 0

    for result in normalized_results:
        cell_line_id = result.get("cell_line_id", "unknown")
        normalized_data = result.get("normalized_data", {})

        logger.info(f"Validating cell line data for {cell_line_id}")

        try:
            # Validate against JSONOutputSchema Pydantic model
            validated_form = JSONOutputSchema(**normalized_data)

            # Create validated result with the validated data
            validated_result = {
                **result,  # Keep original metadata (cell_line_id, processing_times, etc.)
                "validated_data": validated_form.model_dump(),
                "validation_status": "success"
            }

            validated_results.append(validated_result)
            logger.info(f"Successfully validated cell line {cell_line_id}")

        except ValidationError as e:
            validation_errors += 1
            error_msg = f"Validation failed for {cell_line_id}: {str(e)}"
            logger.error(error_msg)

            # Include failed validation in results for debugging
            failed_result = {
                **result,
                "validation_status": "failed",
                "validation_error": str(e),
                "validation_details": e.errors()
            }

            validated_results.append(failed_result)

        except Exception as e:
            validation_errors += 1
            error_msg = f"Unexpected error validating {cell_line_id}: {str(e)}"
            logger.error(error_msg)

            failed_result = {
                **result,
                "validation_status": "error",
                "validation_error": str(e)
            }

            validated_results.append(failed_result)

    successful_validations = len(validated_results) - validation_errors
    logger.info(f"Validation completed: {successful_validations} successful, {validation_errors} failed")

    return validated_results

async def cleanup_and_prepare_result(pdf_info: PDFInfo, validated_results: List[Dict[str, Any]], 
                                   total_start_time: float, cell_lines_found: List[str]) -> Dict[str, Any]:
    """
    Cleanup uploaded files and prepare final result.
    
    Args:
        pdf_info: PDF information including file_id and client
        validated_results: Final validated results
        total_start_time: When the entire process started
        cell_lines_found: Original list of identified cell lines
        
    Returns:
        Final result dictionary
    """
    logger.info("STAGE 8: Cleanup and result preparation...")
    
    # Cleanup uploaded file
    try:
        await pdf_info.client.files.delete(pdf_info.file_id)
        logger.info(f"Cleaned up uploaded file: {pdf_info.file_id}")
    except Exception as e:
        logger.warning(f"Failed to delete uploaded file {pdf_info.file_id}: {e}")
    
    # Calculate total processing time
    total_time = time.time() - total_start_time
    
    # Prepare final result
    final_result = {
        "status": "success",
        "filename": pdf_info.filename,
        "total_processing_time": total_time,
        "cell_lines_found": len(cell_lines_found),
        "successful_validations": len([r for r in validated_results if r.get("validation_status") == "success"]),
        "results": validated_results,
        "identification_result": cell_lines_found
    }
    
    logger.info(f"Final result summary: {final_result['successful_validations']} successful validations out of {final_result['cell_lines_found']} cell lines found")
    logger.info(f"Curation process completed! Total time: {total_time:.2f}s")
    
    return final_result